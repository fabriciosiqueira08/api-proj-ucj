from openpyxl.styles import Font, Alignment
from ProcessCard import process_card
from datetime import datetime

def clear_worksheet(ws):
    for row in ws.iter_rows(min_row=2):  # Começa na segunda linha
        for cell in row:
            cell.value = None

def process_phases(ws, headers, all_phases, row_num):
    clear_worksheet(ws)
    normal_font = Font(name='Arial', size=11, bold=False)
    alignment_bottom = Alignment(vertical='bottom')

    for phase in all_phases:
        if isinstance(phase, dict):
            for card_edge in phase['cards']['edges']:
                card = card_edge['node']
                field_values = process_card(card, headers)

                created_at_str = card.get('createdAt', '')
                if created_at_str:
                    try:
                        created_at = datetime.fromisoformat(created_at_str.replace("Z", "+00:00"))
                        created_at_formatted = created_at.strftime('%d/%m/%Y')
                    except ValueError:
                        created_at_formatted = created_at_str
                else:
                    created_at_formatted = ''

                ws.cell(row=row_num, column=1, value=created_at_formatted)

                for col_num, header in enumerate(headers[1:], 2):
                    cell_value = field_values.get(header, "")
                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    cell.font = normal_font
                    cell.alignment = alignment_bottom

                row_num += 1

    return ws, row_num