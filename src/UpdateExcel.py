import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ProcessPhases import process_phases
from datetime import datetime

def update_excel(wb, all_phases, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    headers = [
        "Criação do Card", "Nome da Empresa", "Nome do Cliente", "Data de Chegada", "Classificação do Cliente",
        "Canal de Chegada", "Data de Atendimento", "Status Conversão para Diagnóstico",
        "Status Conformidade", "Motivo Inconformidade", "Data do Diagnóstico",
        "Status conversão para proposta", "Data de Entrega da Proposta", "Valor oferecido do Projeto",
        "Valor da Hora do Projeto", "Status Orientação Proposta", "Data de Resposta do Cliente",
        "Resposta do Cliente", "Data de Assinatura do Contrato", "Preço Vendido",
        "Preço da Hora Vendida", "Etiqueta Indicação", "Subcanal de Chegada - Indicação", "Projeto é colaborativo?  "
    ]

    # Aplicando os cabeçalhos e seus estilos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.alignment = Alignment(vertical='bottom')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 2
    ws, row_num = process_phases(ws, headers, all_phases, row_num)

    # Extraindo os dados das células para ordenar
    data_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers), values_only=True):
        # Removendo "R$" das células na coluna N e convertendo para número
        cleaned_row = []
        for col_num, cell in enumerate(row, 1):
            if col_num == 14:  # Coluna N é a 14ª coluna
                if isinstance(cell, str) and cell.startswith("R$"):
                    cell = cell.replace("R$", "").strip()
                try:
                    cell = float(cell.replace(",", ".")) if isinstance(cell, str) else cell
                except ValueError:
                    pass
            cleaned_row.append(cell)
        data_rows.append(tuple(cleaned_row))

    # Função para verificar se o valor da coluna N corresponde aos padrões especificados
    def should_remove(value):
        if isinstance(value, str):
            patterns = [
                r'^\d+\s*a\s*\d+$',  
                r'^\d+\s*e\s*\d+$',  
                r'^\d+,\d+$'         
            ]
            for pattern in patterns:
                if re.match(pattern, value):
                    return True
        # Remover exatamente o valor "16,5"
        if isinstance(value, str) and value.strip() == "16,5":
            return True
        return False

    # Filtrando as linhas que não devem ser removidas
    filtered_data_rows = [row for row in data_rows if not should_remove(row[13])]

    # Ordenar os dados pela primeira coluna ("Criação do Card"), em ordem decrescente
    filtered_data_rows.sort(key=lambda x: datetime.strptime(x[0], '%d/%m/%Y') if x[0] else datetime.min, reverse=True)

    # Limpar as linhas antigas
    ws.delete_rows(2, ws.max_row)

    # Escrever os dados ordenados de volta na planilha
    for row in filtered_data_rows:
        ws.append(row)

    # Ajuste das colunas
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

    print(f"Dados atualizados na aba '{sheet_name}'.")
