from FetchAllCards import fetch_all_cards
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from UpdateExcel import update_excel
from UpdateExcelExecução import update_excel_execução
from openpyxl import Workbook, load_workbook
from Definitions import PIPE_IDS, PIPE_TO_FILE
import sys
import threading

class RedirectOutput:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, message):
        self.text_widget.insert(tk.END, message)
        self.text_widget.see(tk.END)

    def flush(self):
        pass

def update_excel_files():
    progress_bar['value'] = 0
    text_log.delete(1.0, tk.END)

    selected_dir = directory.get()
    if not selected_dir:
        messagebox.showwarning("Caminho não selecionado", "Por favor, selecione o caminho onde os arquivos estão localizados.")
        return

    update_functions = {
        'Café de Vendas': update_excel,
        'Execução': update_excel_execução,
    }

    total_pipes = len(PIPE_TO_FILE)
    progress_bar['maximum'] = total_pipes

    for index, (pipe_name, (filename, sheet_name)) in enumerate(PIPE_TO_FILE.items()):
        full_path = f"{selected_dir}/{filename}"
        print(f"Iniciando a consulta dos dados do Pipefy para: {pipe_name}")
        all_phases = fetch_all_cards(PIPE_IDS[pipe_name])

        try:
            wb = load_workbook(full_path)
            print(f"Arquivo '{filename}' carregado com sucesso.")
        except FileNotFoundError:
            wb = Workbook()
            print(f"Arquivo '{filename}' não encontrado, criando novo arquivo.")
            wb.remove(wb.active)  # Remover a aba padrão vazia

        update_function = update_functions[pipe_name]
        update_function(wb, all_phases, sheet_name)

        wb.save(full_path)
        print(f'Arquivo "{filename}" salvo com sucesso com a aba atualizada.')

        progress_bar['value'] = index + 1
        root.update_idletasks()

    messagebox.showinfo("Concluído", "Atualização concluída com sucesso!")

def start_update_thread():
    threading.Thread(target=update_excel_files).start()

def select_directory():
    path = filedialog.askdirectory()
    if path:
        directory.set(path)

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Atualizar Planilhas de PROJ")

    directory = tk.StringVar()

    tk.Label(root, text="Selecione o caminho das planilhas:").pack(pady=10)
    tk.Entry(root, textvariable=directory, width=50).pack(pady=5)
    tk.Button(root, text="Selecionar Caminho", command=select_directory).pack(pady=5)

    progress_bar = ttk.Progressbar(root, orient='horizontal', length=300, mode='determinate')
    progress_bar.pack(pady=10)

    tk.Button(root, text="Atualizar Planilhas", command=start_update_thread).pack(pady=20)

    text_log = tk.Text(root, height=10, wrap='word', bg='white', fg='black')
    text_log.pack(pady=10, padx=10, fill='both', expand=True)

    sys.stdout = RedirectOutput(text_log)

    root.mainloop()
