from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ProcessPhasesExecução import process_phases_execução
from datetime import datetime
import re

def clean_column_d(value):
    if isinstance(value, str):
        return re.sub(r'[^a-zA-Z, ]', '', value)
    return value

def update_excel_execução(wb, all_phases, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    headers = [
        "Criação do Card", "Projeto", "Iniciativa ou Projeto", "Consultores alocados", "Gerente alocado", "Data de início do projeto", 
        "Data esperada de finalização", "Houve saída de consultor?", "Consultor que saiu do projeto", 
        "Data de saída do consultor", "Entrou algum consultor", "Novo consultor", "Data de entrada do novo consultor", 
        "Houve troca de gerente?", "Data da troca de gerente", "Data de início do projeto", "Data esperada de finalização", 
        "Novo gerente", "Houve paralisação", "Data do início da paralisação", "Data de fim da paralisação"
    ]

    # Aplicando os cabeçalhos e seus estilos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.alignment = Alignment(vertical='bottom')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 2
    ws, row_num = process_phases_execução(ws, headers, all_phases, row_num)

    # Extraindo os dados das células para ordenar
    data_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers), values_only=True):
        cleaned_row = []
        for col_num, cell in enumerate(row, 1):
            if col_num == 4:  # Coluna D é a 4ª coluna
                cell = clean_column_d(cell)
            cleaned_row.append(cell)
        data_rows.append(tuple(cleaned_row))
    
    # Ordenar os dados pela primeira coluna ("Projeto"), em ordem decrescente
    data_rows.sort(key=lambda x: datetime.strptime(x[0], '%d/%m/%Y') if x[0] else datetime.min, reverse=True)

    # Limpar as linhas antigas
    ws.delete_rows(2, ws.max_row)

    # Escrever os dados ordenados de volta na planilha
    for row in data_rows:
        ws.append(row)

    # Ajuste das colunas
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

    print(f"Dados atualizados na aba '{sheet_name}'.")
